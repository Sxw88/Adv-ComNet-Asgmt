# importing matplotlib module 
from matplotlib import pyplot as plt

# Num PY
import numpy as np

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

#how many cycles
cycles = 5
#how many measurements per cycle
msmt = 8

# Switcher is dictionary data type here
def numbers_to_alph(argument):
    switcher = {
        0: "A",
        1: "B",
        2: "C",
        3: "D",
        4: "E",
    }
    return switcher.get(argument, "nothing")

# Load in the workbook
wb = load_workbook('Calculations.xlsx')

# Get sheet names
print(wb.sheetnames)

# Get a sheet by name
sheet = wb['Sheet1']

# Print the sheet title
print('Sheet Title:',sheet.title)
  
# x-axis values 
x = list()
  
# Y-axis values 
y = list()

#Reading data from excel sheet and putting them into arrays
for i in range(cycles):
    col = numbers_to_alph(i)
    for j in range(msmt):
        dist = (j+1)*2
        pos = col + str(j+1)
        y.append(sheet[pos].value)
        x.append(dist)

print("length of x is", len(x), "\nx = ", x)
print("length of y is", len(y), "\ny = ", y)

# Adding Title
plt.title("IDK What to put yet\n")

# Labeling the axes
plt.xlabel("Distance (m)")
plt.ylabel("Path Loss (dB)")

#calculating best fit curve using polyfit
# Calculating parameters (theta0, theta1)
# of the log curve using the numpy.polyfit() function
x_div_2 = [k / 2 for k in x]
theta = np.polyfit(np.log10(x_div_2), y, 1)

print(f'The parameters of the curve: {theta}')
print(f'Best fit Log curve equation: Y = ', theta[0], 'log X + ', theta[1])

# Now, calculating the y-axis values against x-values according to
# the parameters theta0, theta1 and theta2
#y_line = theta[2] + theta[1] * pow(np.array(x), 1) + theta[0] * pow(np.array(x), 2)
y_line = list()

for i in x_div_2:
    j = theta[0] * np.log10(i) + theta[1]
    y_line.append(j)

print("length of y_line is", len(y_line), "\ny_line = ", y_line[0:msmt])

# Function to plot
plt.scatter(x, y, color="steelblue")
plt.plot(x[0:msmt], y_line[0:msmt], color="red")

# function to show the plot 
plt.show()
