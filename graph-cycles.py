# importing matplotlib module 
from matplotlib import pyplot as plt

# Num PY
import numpy as np

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

#how many cycles
cycles = 5
#how many measurements per cycle
msmt = 8

# Switcher is dictionary data type here
def numbers_to_alph(argument):
    switcher = {
        0: "A",
        1: "B",
        2: "C",
        3: "D",
        4: "E",
    }
    return switcher.get(argument, "nothing")

#list_x = the list which is nesting other sublists, n = which element to print out
def first_item_in_sublist(list_x, n):
    output_list = list()
    for i in list_x:
        output_list.append(i[n])
    return output_list


# placing plots in the plane
fig, plots = plt.subplots(3,2)

plots[0,0].set_title('cycle 1')
plots[0,1].set_title('cycle 2')
plots[1,0].set_title('cycle 3')
plots[1,1].set_title('cycle 4')
plots[2,0].set_title('cycle 5')

# Labeling the axes
plots[0,0].set_xlabel('Distance (m)')
plots[0,0].set_ylabel('Path Loss (dB)')
plots[0,1].set_xlabel('Distance (m)')
plots[0,1].set_ylabel('Path Loss (dB)')
plots[1,0].set_xlabel('Distance (m)')
plots[1,0].set_ylabel('Path Loss (dB)')
plots[1,1].set_xlabel('Distance (m)')
plots[1,1].set_ylabel('Path Loss (dB)')
plots[2,0].set_xlabel('Distance (m)')
plots[2,0].set_ylabel('Path Loss (dB)')

# Load in the workbook
wb = load_workbook('Calculations.xlsx')

# Get sheet names
print(wb.sheetnames)

# Get a sheet by name
sheet = wb['Sheet1']

# Print the sheet title
print('Excel Sheet Title:',sheet.title)
  
# Creating lists to store all our values
# 3-levels nested lists >> top_list[ cycle_list[ xy_list[x,y] ] ]
top_list = [None] * cycles
cycle_list = list()
xy_list = list()

# Reading data from excel sheet and putting them into arrays
for i in range(cycles):
    col = numbers_to_alph(i)
    cycle_list.clear()
    cycle_list = [None] * msmt
    for j in range(10, msmt+10):
        xy_list.clear()
        dist = (j+1)*2 -20
        pos = col + str(j)
        xy_list.append(dist)                #append x to xy_list
        xy_list.append(sheet[pos].value)    #append y to xy_list
        cycle_list[j-10] = xy_list[:]               #append xy_list to cycle_list
    top_list[i] = cycle_list[:]                         #finally append cycle_list to top_list

print("length of top_list is", len(top_list), "\n\nContents of top_list: ")
for i in top_list:
    print("Length of cycle_list = ", len(cycle_list), "\nContents of cycle_list: \n", i, "\n")

#calculating best fit curve using polyfit
# Calculating parameters (theta0, theta1)
# of the log curve using the numpy.polyfit() function
theta = list()
for i in top_list:
    theta.append(np.polyfit(np.log10(first_item_in_sublist(i,0)), first_item_in_sublist(i,1), 1))

for i in range(cycles):
    print("-------------------------------------------------------------------------------------")
    print(f'The parameters of the curve: {theta[i]}')
    print(f'Best fit Log curve equation: Y = ', theta[i][0], 'log X + ', theta[i][1], "\n")
print("-----------------------------------------------------------------------------------------")

# Now, calculating the y-axis values against x-values according to
# the parameters theta0, theta1
y_line= [None] * cycles
y_line_sublist = list()

for k in top_list:
    y_line_sublist.clear()
    j = theta[top_list.index(k)][0] * np.log10(first_item_in_sublist(k,0)) + theta[top_list.index(k)][1]
    y_line_sublist.append(j)
    y_line[top_list.index(k)] = y_line_sublist[:]

grid_x = 0
grid_y = 0
grid_width = 2

# Function to plot
for i in top_list:
    x = first_item_in_sublist(i,0)
    plots[grid_x,grid_y].scatter(x, first_item_in_sublist(i,1), color="steelblue")
    y_line_tolist = y_line[top_list.index(i)][0:msmt][0].tolist()
    plots[grid_x,grid_y].plot(x[0:msmt], y_line_tolist, color="red")

    grid_y += 1
    if grid_y == grid_width:
        grid_x += 1
        grid_y = 0

# function to show the plot
plt.tight_layout()
plt.show()
